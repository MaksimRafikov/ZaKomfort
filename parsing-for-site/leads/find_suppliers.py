#!/usr/bin/env python3
"""Сбор контактов потенциальных ПОКУПАТЕЛЕЙ базы — поставщиков и сервисов
фитнес-индустрии Москвы/МО (вариант №1: прямые продажи).

Список компаний собран через веб-поиск по нишам. Скрипт заходит на сайт каждой
компании (главная + страницы контактов) и вытаскивает телефон и e-mail, затем
пишет таблицу лидов в CSV и XLSX.

Запуск: python3 find_suppliers.py --workers 12
"""
from __future__ import annotations
import argparse
import csv
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_CSV = os.path.join(HERE, "suppliers_moscow.csv")
OUT_XLSX = os.path.join(HERE, "suppliers_moscow.xlsx")

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
CONTACT_PATHS = ("", "contacts", "contacts/", "kontakty", "kontakty/", "contact",
                 "about", "o-kompanii", "o-nas")

# (Компания, Категория, домен)
SEEDS = [
    ("СпортРес", "Тренажёры/оборудование", "sportres.ru"),
    ("ПРОФИФИТ", "Тренажёры/оборудование", "profi.fit"),
    ("VivaSport", "Тренажёры/оборудование", "vivasport.ru"),
    ("LifeTren", "Тренажёры/оборудование", "lifetren.pro"),
    ("Proftren", "Тренажёры/оборудование", "proftren.ru"),
    ("All4gym", "Свободные веса (производитель)", "all4gym.ru"),
    ("Barbell Atlet", "Свободные веса (производитель)", "barbellatlet.ru"),
    ("Profigym", "Свободные веса (производитель)", "profi-gym.ru"),
    ("Атлант-Спорт", "Свободные веса (производитель)", "atlant-sport.ru"),
    ("Кокленков (гантели Атлант)", "Свободные веса (производитель)", "koklenkov.ru"),
    ("Трамплин-Спорт", "Кроссфит/функц. тренинг", "tramplinsport.ru"),
    ("Форма-Спорта", "Кроссфит/функц. тренинг", "forma-sporta.com"),
    ("YOUSTEEL", "Кроссфит/функц. тренинг", "yousteel.ru"),
    ("Sports-Tech", "Кроссфит/функц. тренинг", "sports-tech.ru"),
    ("Аквабас", "Бассейны: оборудование/химия", "shop.aquabas.ru"),
    ("Бассейн.ру", "Бассейны: оборудование/химия", "basseyn.ru"),
    ("ПулМаркет", "Бассейны: оборудование/химия", "poolmarket.ru"),
    ("ТД ЭКТИС", "Бассейны: оборудование/химия", "ectes-td.ru"),
    ("1С:Фитнес клуб", "CRM/ПО для клубов", "fitness1c.ru"),
    ("FitBase", "CRM/ПО для клубов", "fitbase.io"),
    ("Mobifitness", "CRM/ПО для клубов", "mobifitness.ru"),
    ("4Бит (1С)", "CRM/ПО, касса", "4bit.ru"),
    ("Helix Group", "CRM/ПО для клубов", "helix-group.ru"),
    ("Онлайн-касса.ru", "Кассы/эквайринг", "online-kassa.ru"),
    ("Касса 77", "Кассы/эквайринг", "kassa77.ru"),
    ("ЮНИВЕРС-СОФТ", "СКУД/турникеты", "universe-soft.ru"),
    ("Skudov.net", "СКУД/турникеты", "skudov.net"),
    ("Технологии Успеха", "СКУД/турникеты", "tech-success.ru"),
    ("СКУД-ПРО", "СКУД/турникеты", "pro-skud.ru"),
    ("FitAtletik", "Спортпит опт", "fitatletik-sport.ru"),
    ("Optstrong", "Спортпит опт", "optstrong.ru"),
    ("SportFood", "Спортпит опт", "sportfood40.ru"),
    ("4mass", "Спортпит опт", "4mass.ru"),
    ("Академия-Т", "Спортпит опт", "ac-t.ru"),
    ("REOVAX", "Маркетинг/SMM для фитнеса", "reovax.ru"),
    ("NOVA FIT CONSULTING", "Маркетинг/консалтинг", "novafitconsulting.ru"),
    ("Magnetic agency", "Маркетинг/SMM для фитнеса", "magneticagency.ru"),
    ("Агентство фитнес-маркетинга", "Маркетинг/SMM для фитнеса", "fitness-marketing.ru"),
    ("56PX", "Маркетинг/SMM", "56px.ru"),
    ("TechenGroup", "Напольные покрытия", "techengroup.ru"),
    ("Олимпик Филд", "Напольные покрытия", "olympic-field.ru"),
    ("Хантсман-НМГ (Daltosport)", "Напольные покрытия", "huntsman-nmg.com"),
    ("Резком", "Напольные покрытия", "rezcom.ru"),
    ("БруКлин Групп", "Клининг", "brooclean.ru"),
    ("Гаусс-сервис", "Клининг", "gauss-service.ru"),
    ("Легис Клининг", "Клининг", "legis-cleaning.ru"),
    ("ЭкспертЧистоты", "Клининг", "qlean-m.ru"),
    ("Гармония Чистоты", "Клининг", "harmony-clean.ru"),
    ("MAX Clean Room", "Вентиляция/климат", "maxpv.ru"),
    ("ИНТЕХ", "Вентиляция/климат", "air-ventilation.ru"),
    ("Империя Климата", "Вентиляция/климат", "vrf-mrv.ru"),
    ("Технологии Микроклимата", "Вентиляция/климат", "climate-technology.ru"),
    ("МирВент", "Вентиляция/климат", "mir-vent.ru"),
    ("Barssport", "Спортивная форма/одежда", "barssport-factory.ru"),
    ("CleverCut", "Спортивная форма/одежда", "clevercut.ru"),
    ("Флоренс", "Спортивная форма/одежда", "florens.group"),
    ("Швейная фабрика", "Спортивная форма/одежда", "shveynaya.ru"),
    ("ВКН Системы", "Шкафчики/мебель для раздевалок", "kpshka.ru"),
    ("Мебель-Фитнес", "Шкафчики/мебель для раздевалок", "mebel-fitnes.ru"),
    ("КомплектФит", "Шкафчики/мебель для раздевалок", "komplektfit.ru"),
    ("ШкафКупе", "Шкафчики/мебель для раздевалок", "shkaffkupe.ru"),
    ("СЕАН", "Проектирование/строительство клубов", "sea-n.com"),
    ("СК Созидание", "Проектирование/строительство клубов", "sk-sozidanie.ru"),
    ("Fitness Development", "Проектирование/строительство клубов", "fitnessdevelopment.ru"),
    ("ГК Лидер", "Проектирование/строительство клубов", "l-gk.ru"),
    ("2П Строй", "Проектирование/строительство клубов", "dvapstroi.ru"),
    ("Гелиокс", "SPA/массажное оборудование", "heliox.ru"),
    ("Евро-Спорт", "SPA/массажное оборудование", "evro-sport.ru"),
    ("Old Point", "SPA/массажное оборудование", "oldpoint.ru"),
    ("Мед-Мос", "SPA/массажное оборудование", "med-mos.ru"),
    ("Beauty Technology", "SPA/салонное оборудование", "medispatechnology.ru"),
    ("ГлавСтекло", "Зеркала для залов", "glavglass.ru"),
    ("NAYADA Glass", "Зеркала для залов", "nayada-glass.ru"),
    ("АБС", "Зеркала для залов", "abs-steklo.ru"),
    ("Tanita-shop", "Зеркала для залов", "tanita-shop.ru"),
    ("Европейские стекольные технологии", "Зеркала для залов", "eurogt.ru"),
    ("Экодар", "Вода/кулеры/пурифайеры", "ekodar.ru"),
    ("WW Кулеры", "Вода/кулеры/пурифайеры", "wwkulery.ru"),
    ("Московские фильтры", "Вода/кулеры/пурифайеры", "mosfilters.ru"),
    ("PlastikKarta", "Пластиковые клубные карты", "plastikkarta.ru"),
    ("МосКард", "Пластиковые клубные карты", "moskard.ru"),
    ("Ас Принт", "Пластиковые клубные карты", "asprintcard.ru"),
    ("headup", "Пластиковые клубные карты", "headup.ru"),
    ("Card-online", "Пластиковые клубные карты", "card-online.ru"),
    ("ALFAFIT", "Бокс/единоборства (оборудование)", "alfafit.ru"),
    ("TOTALBOX", "Бокс/единоборства (оборудование)", "totalbox.pro"),
    ("Fighttech", "Бокс/единоборства (оборудование)", "fighttech.ru"),
    ("FILIPPOV DYNASTY", "Бокс/единоборства (оборудование)", "filippov-dynasty.ru"),
    ("Slamix (Euromat)", "Бокс/единоборства (ринги)", "slamix.ru"),
    ("Евромат", "Маты/татами/ковры", "euro-mat.ru"),
    ("ОКТАКЕМ", "Маты/татами/ковры", "oktakem.ru"),
    ("Fire-Sun", "Солярии", "fire-sun.ru"),
    ("Все Для Солярия", "Солярии", "solyarij-luxura.ru"),
    ("Profi-Market", "Солярии", "profi-market.ru"),
    ("Zagar-shop", "Солярии", "zagar-shop.ru"),
    ("Солана", "Солярии", "solana.ru"),
    ("VendLiga", "Вендинг", "vendliga.ru"),
    ("Double Black", "Вендинг/снеки/вода", "dblack.ru"),
    ("Rem Vend", "Вендинг", "remvend.ru"),
    ("QuickUp", "Вендинг (спортпит)", "quickup.ru"),
    ("Мир Вендинга (TCN)", "Вендинг", "tcn.ru"),
    ("i-Motion EMS", "EMS-оборудование", "emskostum.ru"),
    ("Beautec (i-Motion)", "EMS-оборудование", "beautec.ru"),
    ("VIP имидж", "EMS-оборудование", "vipim.ru"),
    ("Vilmed (ЭСМА)", "EMS-оборудование", "vilmed.ru"),
    ("ЭСМА", "EMS-оборудование", "esma.ru"),
    ("БиЭндСи СПА", "Сауны/хамам под ключ", "spabuild.ru"),
    ("Хамам Люкс", "Сауны/хамам под ключ", "hamam-lux.ru"),
    ("Непарим", "Сауны/хамам/бассейны под ключ", "neparim.ru"),
    ("Арт-Хамам", "Сауны/хамам под ключ", "art-hamam.com"),
    ("Интерлайт", "Освещение для залов", "interlight.ru"),
    ("Twin Light", "Освещение для залов", "twinlight.ru"),
    ("Локус Лайт", "Освещение для залов", "locus-light.ru"),
    ("Prof-LED", "Освещение для залов", "prof-led.ru"),
    ("Апекс-энерго", "Освещение для залов", "apex-energy.ru"),
]

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
PHONE_RE = re.compile(r"(?:\+7|8)[\s\-]?\(?\d{3,4}\)?[\s\-]?\d{2,3}[\s\-]?\d{2}[\s\-]?\d{2}")
BAD_EMAIL = ("example.", "sentry.", "wixpress", "@2x", ".png", ".jpg", ".gif",
             ".webp", ".svg", "your@", "email@", "domain.", "test@", "u003e",
             "react", "schema.org", "@sentry", "yandex.ru/clck", "@2x.")


def clean_emails(text):
    out = []
    for e in EMAIL_RE.findall(text):
        el = e.lower()
        if any(b in el for b in BAD_EMAIL):
            continue
        out.append(e)
    pref = [e for e in out if e.lower().split("@")[0] in
            ("info", "mail", "sales", "zakaz", "opt", "office", "client",
             "shop", "manager", "hello", "zayavka")]
    return list(dict.fromkeys(pref + out))


def norm_phone(p):
    d = re.sub(r"\D", "", p)
    if len(d) == 11 and d[0] == "8":
        d = "7" + d[1:]
    if len(d) == 11 and d[0] == "7":
        return f"+7 ({d[1:4]}) {d[4:7]}-{d[7:9]}-{d[9:11]}"
    return p.strip()


def is_junk_phone(np):
    d = re.sub(r"\D", "", np)
    if len(d) != 11:
        return True
    if d[1] == "0":          # коды РФ не начинаются с 0
        return True
    if re.search(r"(\d)\1{6,}", d):   # подряд 7+ одинаковых цифр (плейсхолдеры)
        return True
    return False


def clean_phones(text):
    out = []
    for p in PHONE_RE.findall(text):
        np = norm_phone(p)
        if not is_junk_phone(np):
            out.append(np)
    return list(dict.fromkeys(out))


# Контакты, добитые вручную из результатов поиска (сайт не отдал/закрыт)
MANUAL = {
    "alfafit.ru": ("+7 (495) 984-89-93", "client@alfafit.ru"),
    "tanita-shop.ru": ("+7 (495) 504-51-00", ""),
}


def fetch(session, url):
    try:
        r = session.get(url, timeout=18, headers={"User-Agent": UA}, allow_redirects=True)
        if r.status_code == 200 and r.text:
            return r.text
    except Exception:
        return None
    return None


def process(seed):
    name, category, domain = seed
    phones, emails = [], []
    with requests.Session() as s:
        for path in CONTACT_PATHS:
            for base in ("https://" + domain, "http://" + domain):
                html = fetch(s, base.rstrip("/") + "/" + path)
                if html:
                    phones += clean_phones(html)
                    emails += clean_emails(html)
                    break
            if phones and emails:
                break
    phones = list(dict.fromkeys(phones))[:3]
    emails = list(dict.fromkeys(emails))[:2]
    if domain in MANUAL:
        mp, me = MANUAL[domain]
        if not phones and mp:
            phones = [mp]
        if not emails and me:
            emails = [me]
    return {
        "Компания": name,
        "Категория (что продаём)": category,
        "Сайт": "https://" + domain,
        "Телефон": " ; ".join(phones),
        "E-mail": " ; ".join(emails),
        "Город": "Москва/МО",
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workers", type=int, default=12)
    args = ap.parse_args()

    rows = []
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(process, s): s for s in SEEDS}
        for i, fut in enumerate(as_completed(futs), 1):
            rows.append(fut.result())
            if i % 20 == 0:
                print(f"  {i}/{len(SEEDS)} | {i/(time.time()-t0):.1f}/с", flush=True)

    # оставляем только компании с хотя бы одним контактом
    rows = [r for r in rows if r["Телефон"] or r["E-mail"]]
    rows.sort(key=lambda r: (r["Категория (что продаём)"], r["Компания"]))
    cols = ["Компания", "Категория (что продаём)", "Сайт", "Телефон", "E-mail", "Город"]
    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)

    with_phone = sum(1 for r in rows if r["Телефон"])
    with_email = sum(1 for r in rows if r["E-mail"])
    with_any = sum(1 for r in rows if r["Телефон"] or r["E-mail"])
    print(f"Всего компаний: {len(rows)} | с телефоном: {with_phone} | "
          f"с e-mail: {with_email} | с любым контактом: {with_any}")
    print(f"CSV: {OUT_CSV}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Поставщики Москва"
    ws.append(cols)
    for r in rows:
        ws.append([r[c] for c in cols])
    fill = PatternFill("solid", fgColor="1F4E78")
    for ci in range(1, len(cols) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = fill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(vertical="center")
    for ci, w in zip(range(1, len(cols) + 1), (32, 34, 30, 46, 34, 12)):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
    wb.save(OUT_XLSX)
    print(f"XLSX: {OUT_XLSX}")


if __name__ == "__main__":
    main()
