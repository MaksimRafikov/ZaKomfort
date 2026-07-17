#!/usr/bin/env python3
"""Удобный для пользователя Excel-образец базы (демо).

Делает красивый demo_sample.xlsx из demo_sample.csv:
  - титульный лист «О базе» с описанием и цифрами;
  - лист с таблицей: автофильтр, закреплённая шапка, чередование строк,
    кликабельные ссылки (сайт/соцсети, e-mail, телефон), перенос текста.

Запуск: python3 build_demo_xlsx.py
"""
from __future__ import annotations
import csv
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "data"))
SRC = os.path.join(DATA, "demo_sample.csv")
OUT = os.path.join(DATA, "demo_sample.xlsx")

# Порядок и ширины колонок (по убыванию пользы для покупателя)
COLS = [
    ("Название клуба", 36),
    ("Бренд/сеть", 18),
    ("Тип объекта", 18),
    ("Город", 16),
    ("Адрес", 42),
    ("Телефон", 22),
    ("E-mail", 26),
    ("Режим работы", 34),
    ("Ссылки на сайт или соц. сети", 40),
    ("Дата снимка", 13),
]

NAVY = "1F4E78"
LIGHT = "EAF1FB"
LINKBLUE = "0563C1"


def first_url(text: str) -> str:
    for part in re.split(r"[;\s]+", text or ""):
        if part.startswith("http"):
            return part
    return ""


def main():
    rows = list(csv.DictReader(open(SRC, encoding="utf-8-sig")))
    headers = [c[0] for c in COLS]

    wb = Workbook()

    # ---------- Лист 1: О базе ----------
    cover = wb.active
    cover.title = "О базе"
    cover.sheet_view.showGridLines = False
    cover.column_dimensions["A"].width = 2
    cover.column_dimensions["B"].width = 100

    def put(row, text, size=11, bold=False, color="1F2A37"):
        c = cover.cell(row=row, column=2, value=text)
        c.font = Font(size=size, bold=bold, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        return c

    cover.cell(row=2, column=2, value="База спортивных объектов России — ДЕМО")
    cover.cell(row=2, column=2).font = Font(size=20, bold=True, color=NAVY)
    put(4, "Это бесплатный демонстрационный образец (60 клубов из 22 городов). "
           "Полная база содержит 6 632 объекта по всей России.", 12, bold=True)
    put(6, "Что в каждой строке:", 12, bold=True)
    put(7, "•  Название клуба, Бренд/сеть, Тип объекта (фитнес-клуб / бассейн / "
           "спорткомплекс / студия)")
    put(8, "•  Город, Адрес")
    put(9, "•  Телефон, E-mail")
    put(10, "•  Режим работы")
    put(11, "•  Ссылки на сайт и соцсети")
    put(13, "Полная версия — заполненность полей:", 12, bold=True)
    put(14, "Название/Город/Адрес/Режим — 100% • Телефон — 97% • E-mail — 41% • "
            "Ссылки — 93%")
    put(16, "Форматы полной поставки:", 12, bold=True)
    put(17, "•  Единый CSV и Excel (с фильтрами и сводками по городам/типам/сетям)")
    put(18, "•  Разбивка по отдельному файлу на каждый город")
    put(19, "•  Срезы под нишу по запросу (например, «только бассейны»)")
    put(21, "Перейдите на лист «Демо (образец)» внизу окна, чтобы посмотреть данные.",
        11, bold=True, color=NAVY)

    # ---------- Лист 2: данные ----------
    ws = wb.create_sheet("Демо (образец)")
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # шапка
    head_fill = PatternFill("solid", fgColor=NAVY)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci)
        c.fill = head_fill
        c.font = Font(color="FFFFFF", bold=True, size=11)
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # ширины + перенос для длинных полей
    wrap_cols = {"Адрес", "Режим работы", "Ссылки на сайт или соц. сети"}
    for ci, (name, width) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    idx = {h: i + 1 for i, h in enumerate(headers)}
    link_font = Font(color=LINKBLUE, underline="single")
    for ri in range(2, len(rows) + 2):
        for name in wrap_cols:
            ws.cell(row=ri, column=idx[name]).alignment = Alignment(
                wrap_text=True, vertical="top")
        for name in headers:
            if name not in wrap_cols:
                ws.cell(row=ri, column=idx[name]).alignment = Alignment(vertical="top")
        # кликабельный e-mail
        em = ws.cell(row=ri, column=idx["E-mail"])
        if em.value and "@" in str(em.value) and ";" not in str(em.value):
            em.hyperlink = "mailto:" + str(em.value).strip()
            em.font = link_font
        # кликабельная ссылка (первая из списка)
        lk = ws.cell(row=ri, column=idx["Ссылки на сайт или соц. сети"])
        url = first_url(str(lk.value))
        if url:
            lk.hyperlink = url
            lk.font = link_font

    # таблица с чередованием строк + автофильтр
    ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    table = Table(displayName="DemoTable", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    wb.save(OUT)
    print(f"Готово: {OUT} ({len(rows)} строк)")


if __name__ == "__main__":
    main()
