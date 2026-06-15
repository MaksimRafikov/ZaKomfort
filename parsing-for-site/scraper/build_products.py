#!/usr/bin/env python3
"""Сборка продаваемых артефактов из основного CSV.

Создаёт:
  data/sportgyms_clubs.xlsx     — оформленный XLSX (фильтры, закреплённая шапка,
                                   листы «Города», «Типы», «Сети»)
  data/by_city/<Город>.csv      — отдельный файл на каждый город
  data/demo_sample.csv          — бесплатный демо-образец (~60 строк)

Запуск: python3 build_products.py
"""
from __future__ import annotations
import collections
import csv
import os
import re

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "data"))
MAIN = os.path.join(DATA, "sportgyms_clubs.csv")
XLSX = os.path.join(DATA, "sportgyms_clubs.xlsx")
BYCITY = os.path.join(DATA, "by_city")
DEMO = os.path.join(DATA, "demo_sample.csv")


def safe_name(s: str) -> str:
    s = re.sub(r"[\\/:*?\"<>|]+", "_", s).strip()
    return s or "_"


def main():
    rows = list(csv.DictReader(open(MAIN, encoding="utf-8-sig")))
    cols = list(rows[0].keys())

    # --- per-city CSV ---
    os.makedirs(BYCITY, exist_ok=True)
    for fn in os.listdir(BYCITY):
        if fn.endswith(".csv"):
            os.remove(os.path.join(BYCITY, fn))
    by_city = collections.defaultdict(list)
    for r in rows:
        by_city[r["Город"] or "Без города"].append(r)
    for city, items in by_city.items():
        with open(os.path.join(BYCITY, safe_name(city) + ".csv"),
                  "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            w.writerows(items)
    print(f"Городов-файлов: {len(by_city)} -> {BYCITY}")

    # --- demo sample (до 3 клубов из 22 крупнейших городов) ---
    top_cities = [c for c, _ in collections.Counter(r["Город"] for r in rows).most_common(22)]
    demo = []
    for c in top_cities:
        demo += by_city[c][:3]
    with open(DEMO, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(demo[:60])
    print(f"Демо-образец: {min(len(demo),60)} строк -> {DEMO}")

    # --- XLSX ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl не установлен — XLSX пропущен")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Клубы"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    for ci, _ in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    widths = {
        "Название клуба": 34, "Бренд/сеть": 18, "Тип объекта": 18, "Город": 18,
        "Адрес": 44, "Телефон": 24, "E-mail": 26, "Режим работы": 38,
        "Ссылки на сайт или соц. сети": 46, "URL источника": 44, "Дата снимка": 12,
    }
    for ci, name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, 20)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

    def summary(title, counter, headers):
        s = wb.create_sheet(title)
        s.append(headers)
        for c in (1, 2):
            s.cell(row=1, column=c).fill = header_fill
            s.cell(row=1, column=c).font = header_font
        for k, v in counter.most_common():
            s.append([k, v])
        s.column_dimensions["A"].width = 30
        s.column_dimensions["B"].width = 14
        s.freeze_panes = "A2"

    summary("Города", collections.Counter(r["Город"] for r in rows), ["Город", "Клубов"])
    summary("Типы", collections.Counter(r["Тип объекта"] for r in rows), ["Тип объекта", "Клубов"])
    summary("Сети", collections.Counter(r["Бренд/сеть"] for r in rows if r["Бренд/сеть"]),
            ["Бренд/сеть", "Клубов"])
    wb.save(XLSX)
    print(f"XLSX: {XLSX}")


if __name__ == "__main__":
    main()
